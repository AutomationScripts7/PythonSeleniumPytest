import openpyxl

#Loading the Excel file
excel = openpyxl.load_workbook("/Excel_Validation.xlsx")

#Getting active sheet from Excel
activeSheet = excel.active

# #Getting value from specific cell
print(activeSheet.cell(row=1, column=2).value)
# print(cell.value) --commented

#Accessing through column name
print(activeSheet["A1"].value)

#Writing data into Excel
data_written_in_Excel = activeSheet.cell(row=2, column=2).value = "Provide"
print(data_written_in_Excel)

#Getting maximum row and column in sheet/Excel
print(activeSheet.max_row)
print(activeSheet.max_column)



#for loop for getting all values in Excel
for i in range(1,activeSheet.max_row + 1):
    for j in range(1, activeSheet.max_column + 1):
        print(activeSheet.cell(row =i, column=j).value)

# #I want only SC3 details alone we required
for i in range(1,activeSheet.max_row + 1):
    if activeSheet.cell(row=i, column=1).value == "SC3":
        print("SC3 found")
        for j in range(2, activeSheet.max_column + 1):
            print(activeSheet.cell(row =i, column=j).value)



#Getting value in dictionary
#Expected dict is {"firstname":"Rahulshetty" , "lastname":"shetty" , "email":"rahulshetty@gmail.com"}
Dict = {}
for i in range(1,activeSheet.max_row + 1):
    if activeSheet.cell(row=i, column=1).value == "SC3":        #to get rows
        print("SC3 found")

        for j in range(2, activeSheet.max_column + 1):          #to get columns

            # Dict["S.NO"] = "Provide" - reference how dictionary

            Dict[activeSheet.cell(row=1, column=j).value] = activeSheet.cell(row=i, column=j).value
print(Dict)
#Usualy we update or get value like this in dictionary - Dict["firstname"] = "Rahulshetty".
#but here already values in Excel so we're trying to get that









