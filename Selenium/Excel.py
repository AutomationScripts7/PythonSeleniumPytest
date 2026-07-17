import openpyxl
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

driver = webdriver.Edge()
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.maximize_window()
driver.implicitly_wait(2)

Fruit = "Apple"
Get_text = driver.find_element(By.XPATH, "(//div[text()='"+Fruit+"']//following::div/following::div/div)[1]").text
print(Get_text)
driver.find_element(By.CSS_SELECTOR, "#downloadButton").click()
time.sleep(5)

#Excel phase
Dict = {}
excel_path = "C:\\Users\\615074106\\Downloads\\download.xlsx"
excel_book = openpyxl.load_workbook(excel_path)
sheet = excel_book.active

#Finding row of required fruit
for i in range(1, sheet.max_row+1):
    for j in range(1, sheet.max_column+1):
        if sheet.cell(row=i, column=j).value == Fruit:
            Dict["row"] = i

#Finding column of required field
for i in range(1, sheet.max_column+1):
    if sheet.cell(row=1, column=i).value == "price":
        Dict["column"] = i

#Updating value in excel
sheet.cell(row=Dict["row"], column=Dict["column"]).value = "100"

#Saving the sheet
excel_book.save(excel_path)

#Uploading the Excel
uploadFile = driver.find_element(By.CSS_SELECTOR, "#fileinput")
uploadFile.send_keys(excel_path)
driver.implicitly_wait(5)
wait = WebDriverWait(driver, 5)
wait.until(EC.presence_of_element_located((By.XPATH, "//div[text()='Updated Excel Data Successfully.']")))